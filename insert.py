import argparse
import shutil
import sys
from datetime import datetime

import openpyxl
import pandas as pd

XLSX_PATH_DEFAULT = "Internal purchases.xlsx"
SHEET_NAME = "Накладные на отпуск зап. на ст."
FIRST_DATA_ROW = 3
FIRST_COL = 2  # column B; columns A and M+ are outside the table

COLUMNS = [
    "№", "Наименование, характеристика", "Номенклатурный номер",
    "Единица измерения", "Количество", "Цена за единицу без НДС",
    "Сумма без НДС", "Наименование поставщика", "ИИН/БИН",
    "Номер документа", "Дата документа",
]
NUMERIC_COLUMNS = {"№", "Количество", "Цена за единицу без НДС", "Сумма без НДС"}
LAST_COL = FIRST_COL + len(COLUMNS) - 1


def to_cell_value(column: str, raw: str):
    if raw is None or raw == "":
        return None
    if column in NUMERIC_COLUMNS:
        try:
            value = float(raw)
            return int(value) if value == int(value) else value
        except ValueError:
            return raw
    return raw


def find_next_empty_row(ws, start_row: int = FIRST_DATA_ROW) -> int:
    row = start_row
    while any(ws.cell(row=row, column=c).value is not None for c in range(FIRST_COL, LAST_COL + 1)):
        row += 1
    return row


def main():
    parser = argparse.ArgumentParser(
        description="Append rows from a накладная extraction CSV into Internal purchases.xlsx"
    )
    parser.add_argument("csv_path", help="CSV file to insert (e.g. output.csv from extract_nakladnaya.py)")
    parser.add_argument("-x", "--xlsx", default=XLSX_PATH_DEFAULT, help="Target xlsx file")
    args = parser.parse_args()

    df = pd.read_csv(args.csv_path, dtype=str, keep_default_na=False)
    missing = [c for c in COLUMNS if c not in df.columns]
    if missing:
        print(f"CSV is missing expected column(s): {', '.join(missing)}", file=sys.stderr)
        sys.exit(1)
    if df.empty:
        print("CSV has no data rows — nothing to insert.", file=sys.stderr)
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(args.xlsx)
    except FileNotFoundError:
        print(f"Could not find {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    backup_path = f"{args.xlsx}.{datetime.now():%Y%m%d-%H%M%S}.bak"
    shutil.copy2(args.xlsx, backup_path)

    if SHEET_NAME not in wb.sheetnames:
        print(
            f"Sheet '{SHEET_NAME}' not found in {args.xlsx}. Available sheets: {', '.join(wb.sheetnames)}",
            file=sys.stderr,
        )
        sys.exit(1)
    ws = wb[SHEET_NAME]

    start_row = find_next_empty_row(ws)
    row = start_row
    for _, csv_row in df.iterrows():
        for i, column in enumerate(COLUMNS):
            ws.cell(row=row, column=FIRST_COL + i).value = to_cell_value(column, csv_row[column])
        row += 1

    try:
        wb.save(args.xlsx)
    except PermissionError:
        print(f"Could not save {args.xlsx} — is it open in Excel? Close it and try again.", file=sys.stderr)
        sys.exit(1)

    print(f"Inserted {len(df)} row(s) into '{SHEET_NAME}', rows {start_row}-{row - 1}.")
    print(f"Backup of the previous file saved to {backup_path}")


if __name__ == "__main__":
    main()
